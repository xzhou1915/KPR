import openpyxl

FILE = "sample.xlsx"

wb = openpyxl.load_workbook(FILE, data_only=True)

# --- Tenor nodes ---
ws_rate = wb["ForwardRate"]
rows_rate = [r for r in ws_rate.iter_rows(values_only=True)][1:]  # skip header
nodes = [(row[2], row[1]) for row in rows_rate if row[2] is not None]
node_days  = [n[0] for n in nodes]
node_names = []
for _, ticker in nodes:
    name = ticker.replace(" Curncy", "").replace("CNH", "")
    node_names.append(name if name else "Spot")

# --- Exposure ---
ws_exp = wb["Exposure"]
rows_exp = list(ws_exp.iter_rows(values_only=True))[1:]  # skip header

spot_delta   = 0.0
node_exp     = {name: 0.0 for name in node_names}

for date, value_date, usd_exp in rows_exp:
    if usd_exp is None:
        continue

    days = (value_date - date).days
    spot_delta += usd_exp

    # Clamp beyond last node
    if days >= node_days[-1]:
        node_exp[node_names[-1]] += usd_exp
        continue

    # At or before spot node
    if days <= node_days[0]:
        node_exp[node_names[0]] += usd_exp
        continue

    # Interpolate between surrounding nodes
    for i in range(len(nodes) - 1):
        lo, hi = node_days[i], node_days[i + 1]
        if lo <= days <= hi:
            w_hi = (days - lo) / (hi - lo)
            node_exp[node_names[i]]     += usd_exp * (1 - w_hi)
            node_exp[node_names[i + 1]] += usd_exp * w_hi
            break

# --- Segment exposures via reverse cumulative sum ---
# segment[A→B] = sum of node_exp at B and all nodes beyond B
# because those trades are all exposed to the A→B segment of the curve
segment_exp = {}
running = 0.0
for i in range(len(nodes) - 1, 0, -1):
    running += node_exp[node_names[i]]
    segment_exp[f"{node_names[i-1]}-{node_names[i]}"] = running

ordered = list(reversed(list(segment_exp.items())))

# --- Output ---
W = 46
print("=" * W)
print("  USDCNH Forward Portfolio Risk")
print("=" * W)
print(f"\n  Spot Delta:  ${spot_delta:>15,.0f}\n")
print("  Points Risk (USD notional by segment):")
print(f"  {'Segment':<14} {'Net Exposure':>15}")
print(f"  {'-'*14} {'-'*15}")
for label, exp in ordered:
    print(f"  {label:<14} ${exp:>14,.0f}")
print("=" * W)
