from __future__ import annotations

import argparse
import html
import logging
import math
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from tvDatafeed import Interval, TvDatafeed

from currency_universe import EXTRA_INSTRUMENTS, USD_CURRENCIES


DEFAULT_EXCHANGE = "FX_IDC"
DEFAULT_BARS = 250
DEFAULT_ST_WINDOW = 30
DEFAULT_LT_WINDOW = 180
OUTPUT_DIR = Path("output")
DXY_SYMBOL = "DXY"

logging.getLogger("tvDatafeed.main").setLevel(logging.CRITICAL)


@dataclass(frozen=True)
class FetchResult:
    normalized_symbol: str
    tradingview_symbol: str
    exchange: str
    orientation: str
    prices: pd.Series


def make_tv_client() -> TvDatafeed:
    username = os.getenv("TV_USERNAME")
    password = os.getenv("TV_PASSWORD")
    if username and password:
        return TvDatafeed(username=username, password=password)
    return TvDatafeed()


def clean_hist(hist: pd.DataFrame | None) -> pd.DataFrame | None:
    if hist is None or hist.empty or "close" not in hist.columns:
        return None
    out = hist.copy()
    out.index = pd.to_datetime(out.index).tz_localize(None).normalize()
    out = out[~out.index.duplicated(keep="last")].sort_index()
    out["close"] = pd.to_numeric(out["close"], errors="coerce")
    out = out.dropna(subset=["close"])
    return out if not out.empty else None


def get_hist_safe(
    tv: TvDatafeed,
    symbol: str,
    exchange: str,
    n_bars: int,
) -> pd.DataFrame | None:
    try:
        hist = tv.get_hist(
            symbol=symbol,
            exchange=exchange,
            interval=Interval.in_daily,
            n_bars=n_bars,
        )
    except Exception as exc:
        print(f"failed {exchange}:{symbol}: {exc}")
        return None
    return clean_hist(hist)


def fetch_currency(
    tv: TvDatafeed,
    currency: str,
    exchange: str,
    n_bars: int,
) -> FetchResult | None:
    direct_symbol = f"{currency}USD"
    inverse_symbol = f"USD{currency}"

    direct = get_hist_safe(tv, direct_symbol, exchange, n_bars)
    if direct is not None:
        return FetchResult(
            normalized_symbol=direct_symbol,
            tradingview_symbol=direct_symbol,
            exchange=exchange,
            orientation="direct",
            prices=direct["close"].rename(direct_symbol),
        )

    inverse = get_hist_safe(tv, inverse_symbol, exchange, n_bars)
    if inverse is not None:
        normalized_prices = (1.0 / inverse["close"]).rename(direct_symbol)
        return FetchResult(
            normalized_symbol=direct_symbol,
            tradingview_symbol=inverse_symbol,
            exchange=exchange,
            orientation="inverted",
            prices=normalized_prices,
        )

    return None


def fetch_extra_instrument(
    tv: TvDatafeed,
    instrument: dict[str, str],
    n_bars: int,
) -> FetchResult | None:
    normalized_symbol = instrument["normalized_symbol"]
    tradingview_symbol = instrument["tradingview_symbol"]
    exchange = instrument["exchange"]

    hist = get_hist_safe(tv, tradingview_symbol, exchange, n_bars)
    if hist is None:
        return None

    return FetchResult(
        normalized_symbol=normalized_symbol,
        tradingview_symbol=tradingview_symbol,
        exchange=exchange,
        orientation="direct",
        prices=hist["close"].rename(normalized_symbol),
    )


def fetch_all(
    currencies: Iterable[str],
    extra_instruments: Iterable[dict[str, str]],
    exchange: str,
    n_bars: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    tv = make_tv_client()
    results: list[FetchResult] = []
    failures: list[dict[str, str]] = []

    for currency in currencies:
        print(f"fetching {currency}USD")
        result = fetch_currency(tv, currency, exchange, n_bars)
        if result is None:
            failures.append(
                {
                    "normalized_symbol": f"{currency}USD",
                    "tried_symbols": f"{currency}USD, USD{currency}",
                    "exchange": exchange,
                    "status": "not available",
                }
            )
            continue
        results.append(result)

    for instrument in extra_instruments:
        label = instrument["normalized_symbol"]
        print(f"fetching {label}")
        result = fetch_extra_instrument(tv, instrument, n_bars)
        if result is None:
            failures.append(
                {
                    "normalized_symbol": label,
                    "tried_symbols": instrument["tradingview_symbol"],
                    "exchange": instrument["exchange"],
                    "status": "not available",
                }
            )
            continue
        results.append(result)

    if not results:
        raise RuntimeError("No TradingView FX symbols were retrieved.")

    prices = pd.concat([result.prices for result in results], axis=1, sort=True).sort_index()
    prices = prices.dropna(axis=1, how="all")

    metadata = [
        {
            "normalized_symbol": result.normalized_symbol,
            "tradingview_symbol": result.tradingview_symbol,
            "exchange": result.exchange,
            "orientation": result.orientation,
            "first_date": result.prices.index.min().date().isoformat(),
            "last_date": result.prices.index.max().date().isoformat(),
            "observations": int(result.prices.count()),
        }
        for result in results
    ]
    metadata.extend(failures)
    return prices, pd.DataFrame(metadata)


def calculate_log_returns(prices: pd.DataFrame) -> pd.DataFrame:
    prices = prices.sort_index().ffill()
    log_returns = np.log(prices / prices.shift(1))
    return log_returns.replace([np.inf, -np.inf], np.nan).dropna(how="all")


def calculate_window_correlation(log_returns: pd.DataFrame, window: int) -> pd.DataFrame:
    if len(log_returns) < window:
        raise ValueError(
            f"Need at least {window} return rows, but only {len(log_returns)} are available."
        )
    return log_returns.tail(window).corr().sort_index().sort_index(axis=1)


def order_symbols(symbols: Iterable[str]) -> list[str]:
    ordered = sorted(symbols)
    if DXY_SYMBOL in ordered:
        ordered.remove(DXY_SYMBOL)
        ordered.insert(0, DXY_SYMBOL)
    return ordered


def order_frame_columns(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.loc[:, order_symbols(frame.columns)]


def order_matrix(matrix: pd.DataFrame) -> pd.DataFrame:
    ordered = order_symbols(matrix.columns)
    return matrix.loc[ordered, ordered]


def excel_safe_sheet_name(name: str) -> str:
    return name[:31]


def write_excel(
    path: Path,
    prices: pd.DataFrame,
    log_returns: pd.DataFrame,
    corr_st: pd.DataFrame,
    corr_lt: pd.DataFrame,
    corr_diff: pd.DataFrame,
    metadata: pd.DataFrame,
    st_window: int,
    lt_window: int,
) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        prices.to_excel(writer, sheet_name=excel_safe_sheet_name("prices_XXXUSD"))
        log_returns.to_excel(writer, sheet_name=excel_safe_sheet_name("log_returns"))
        corr_st.to_excel(writer, sheet_name=excel_safe_sheet_name(f"corr_ST_{st_window}D"))
        corr_lt.to_excel(writer, sheet_name=excel_safe_sheet_name(f"corr_LT_{lt_window}D"))
        corr_diff.to_excel(writer, sheet_name=excel_safe_sheet_name("corr_diff_ST_minus_LT"))
        metadata.to_excel(writer, sheet_name=excel_safe_sheet_name("metadata"), index=False)

        workbook = writer.book
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "B2"
            sheet.sheet_view.showGridLines = False
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E8EDF3")
                cell.alignment = Alignment(horizontal="center")
            for column_cells in sheet.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 11), 26)

        for sheet_name in [
            excel_safe_sheet_name(f"corr_ST_{st_window}D"),
            excel_safe_sheet_name(f"corr_LT_{lt_window}D"),
        ]:
            corr_sheet = workbook[sheet_name]
            end_row = corr_sheet.max_row
            end_col = corr_sheet.max_column
            if end_row >= 2 and end_col >= 2:
                corr_sheet.conditional_formatting.add(
                    f"B2:{corr_sheet.cell(end_row, end_col).coordinate}",
                    ColorScaleRule(
                        start_type="num",
                        start_value=-1,
                        start_color="B2182B",
                        mid_type="num",
                        mid_value=0,
                        mid_color="F7F7F7",
                        end_type="num",
                        end_value=1,
                        end_color="2166AC",
                    ),
                )

        diff_sheet = workbook[excel_safe_sheet_name("corr_diff_ST_minus_LT")]
        end_row = diff_sheet.max_row
        end_col = diff_sheet.max_column
        if end_row >= 2 and end_col >= 2:
            diff_sheet.conditional_formatting.add(
                f"B2:{diff_sheet.cell(end_row, end_col).coordinate}",
                ColorScaleRule(
                    start_type="num",
                    start_value=-0.5,
                    start_color="B2182B",
                    mid_type="num",
                    mid_value=0,
                    mid_color="F7F7F7",
                    end_type="num",
                    end_value=0.5,
                    end_color="2166AC",
                ),
            )


def color_for_value(value: float | None, scale: float) -> str:
    if value is None or math.isnan(value):
        return "#f3f4f6"
    value = max(-scale, min(scale, value)) / scale
    if value >= 0:
        intensity = int(255 - value * 120)
        return f"rgb({intensity},{min(255, intensity + 10)},255)"
    intensity = int(255 + value * 120)
    return f"rgb(255,{intensity},{intensity})"


def write_html(
    path: Path,
    prices: pd.DataFrame,
    log_returns: pd.DataFrame,
    corr_st: pd.DataFrame,
    corr_lt: pd.DataFrame,
    corr_diff: pd.DataFrame,
    metadata: pd.DataFrame,
    st_window: int,
    lt_window: int,
) -> None:
    symbols = order_symbols(corr_diff.columns)
    start_date = prices.index.min().date().isoformat()
    end_date = prices.index.max().date().isoformat()

    def matrix_rows(matrix: pd.DataFrame, scale: float, title_kind: str) -> str:
        rows = []
        for row_symbol in symbols:
            cells = [f"<th>{html.escape(row_symbol)}</th>"]
            for col_symbol in symbols:
                raw = matrix.loc[row_symbol, col_symbol]
                value = None if pd.isna(raw) else float(raw)
                display = "" if value is None else f"{value:.2f}"
                if title_kind == "diff":
                    st_value = corr_st.loc[row_symbol, col_symbol]
                    lt_value = corr_lt.loc[row_symbol, col_symbol]
                    st_display = "" if pd.isna(st_value) else f"{float(st_value):.2f}"
                    lt_display = "" if pd.isna(lt_value) else f"{float(lt_value):.2f}"
                    title = (
                        f"{row_symbol} vs {col_symbol}: "
                        f"ST {st_window}D {st_display}, "
                        f"LT {lt_window}D {lt_display}, "
                        f"Diff {display}"
                    )
                else:
                    title = f"{row_symbol} vs {col_symbol} {title_kind}: {display}"
                cells.append(
                    "<td "
                    f"style=\"background:{color_for_value(value, scale)}\" "
                    f"title=\"{html.escape(title)}\">"
                    f"{display}</td>"
                )
            rows.append(f"<tr>{''.join(cells)}</tr>")
        return "".join(rows)

    table_header = "".join(
        f'<th class="sortable" data-sort-col="{index + 1}">{html.escape(symbol)}</th>'
        for index, symbol in enumerate(symbols)
    )
    diff_rows = matrix_rows(corr_diff, 0.5, "diff")
    st_rows = matrix_rows(corr_st, 1.0, f"{st_window}D correlation")
    lt_rows = matrix_rows(corr_lt, 1.0, f"{lt_window}D correlation")

    html_text = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>USD FX Correlation Dashboard</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f8fafc;
      --panel: #ffffff;
      --text: #172033;
      --muted: #657287;
      --line: #d8dee8;
      --accent: #0f766e;
      --accent-2: #b45309;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", Arial, sans-serif;
      background: var(--bg);
      color: var(--text);
    }}
    header {{
      padding: 24px 28px 14px;
      border-bottom: 1px solid var(--line);
      background: #ffffff;
    }}
    h1 {{
      margin: 0 0 8px;
      font-size: 26px;
      letter-spacing: 0;
    }}
    .subtle {{ color: var(--muted); }}
    main {{ padding: 22px 28px 36px; }}
    section {{
      margin-top: 20px;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: hidden;
    }}
    section h2 {{
      margin: 0;
      padding: 14px 16px;
      font-size: 17px;
      border-bottom: 1px solid var(--line);
    }}
    .scroll {{ overflow: auto; max-height: 72vh; }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }}
    th, td {{
      border: 1px solid var(--line);
      padding: 7px 8px;
      text-align: right;
      white-space: nowrap;
    }}
    th {{
      position: sticky;
      top: 0;
      z-index: 1;
      background: #eef2f7;
      color: #243047;
      font-weight: 650;
    }}
    th:first-child {{
      left: 0;
      z-index: 2;
      text-align: left;
    }}
    th.sortable {{
      cursor: pointer;
      user-select: none;
    }}
    th.sortable::after {{
      content: "";
      display: inline-block;
      margin-left: 5px;
      width: 0;
      height: 0;
      vertical-align: middle;
    }}
    th.sortable.sort-asc::after {{
      border-left: 4px solid transparent;
      border-right: 4px solid transparent;
      border-bottom: 6px solid var(--muted);
    }}
    th.sortable.sort-desc::after {{
      border-left: 4px solid transparent;
      border-right: 4px solid transparent;
      border-top: 6px solid var(--muted);
    }}
    td:first-child {{ text-align: left; }}
    .split {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
      gap: 16px;
    }}
    .split section {{ margin-top: 0; }}
    .note {{
      padding: 12px 16px;
      color: var(--muted);
      font-size: 13px;
      border-top: 1px solid var(--line);
    }}
    .tabs {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      margin-bottom: 12px;
    }}
    .tab {{
      border: 1px solid var(--line);
      background: #ffffff;
      color: var(--text);
      border-radius: 6px;
      padding: 8px 12px;
      cursor: pointer;
      font: inherit;
    }}
    .tab.active {{
      background: var(--accent);
      border-color: var(--accent);
      color: #ffffff;
    }}
    .matrix-panel {{ display: none; }}
    .matrix-panel.active {{ display: block; }}
  </style>
</head>
<body>
  <header>
    <h1>USD FX Correlation Dashboard</h1>
    <div class="subtle">Daily log-return correlations, normalized as XXXUSD, {start_date} to {end_date}</div>
  </header>
  <main>
    <div class="tabs" aria-label="Matrix views">
      <button class="tab active" data-target="diff">ST - LT Diff</button>
      <button class="tab" data-target="st">ST {st_window}D</button>
      <button class="tab" data-target="lt">LT {lt_window}D</button>
    </div>

    <section>
      <h2>ST - LT Divergence Matrix</h2>
      <div id="panel-diff" class="matrix-panel active">
        <div class="scroll">
          <table>
            <thead><tr><th>Symbol</th>{table_header}</tr></thead>
            <tbody>{diff_rows}</tbody>
          </table>
        </div>
        <div class="note">Blue means short-term correlation is higher than long-term correlation. Red means short-term correlation is lower. Color scale is capped at +/-0.50.</div>
      </div>
      <div id="panel-st" class="matrix-panel">
        <div class="scroll">
          <table>
            <thead><tr><th>Symbol</th>{table_header}</tr></thead>
            <tbody>{st_rows}</tbody>
          </table>
        </div>
        <div class="note">Short-term correlation uses the latest {st_window} daily log-return observations.</div>
      </div>
      <div id="panel-lt" class="matrix-panel">
        <div class="scroll">
          <table>
            <thead><tr><th>Symbol</th>{table_header}</tr></thead>
            <tbody>{lt_rows}</tbody>
          </table>
        </div>
        <div class="note">Long-term correlation uses the latest {lt_window} daily log-return observations.</div>
      </div>
    </section>

    <section>
      <h2>Method</h2>
      <div class="note">All prices are normalized as XXXUSD. Correlations are Pearson correlations on daily log returns. The divergence matrix is {st_window}D correlation minus {lt_window}D correlation.</div>
    </section>
  </main>
  <script>
    const tabs = document.querySelectorAll(".tab");
    const panels = document.querySelectorAll(".matrix-panel");
    const title = document.querySelector("section h2");
    const labels = {{
      diff: "ST - LT Divergence Matrix",
      st: "ST {st_window}D Correlation Matrix",
      lt: "LT {lt_window}D Correlation Matrix"
    }};
    tabs.forEach((tab) => {{
      tab.addEventListener("click", () => {{
        tabs.forEach((item) => item.classList.remove("active"));
        panels.forEach((panel) => panel.classList.remove("active"));
        tab.classList.add("active");
        document.querySelector("#panel-" + tab.dataset.target).classList.add("active");
        title.textContent = labels[tab.dataset.target];
      }});
    }});

    document.querySelectorAll("th.sortable").forEach((header) => {{
      header.addEventListener("click", () => {{
        const table = header.closest("table");
        const tbody = table.querySelector("tbody");
        const colIndex = Number(header.dataset.sortCol);
        const nextDir = header.classList.contains("sort-desc") ? "asc" : "desc";

        table.querySelectorAll("th.sortable").forEach((item) => {{
          item.classList.remove("sort-asc", "sort-desc");
        }});
        header.classList.add(nextDir === "asc" ? "sort-asc" : "sort-desc");

        const rows = Array.from(tbody.querySelectorAll("tr"));
        rows.sort((left, right) => {{
          const leftValue = Number.parseFloat(left.cells[colIndex].textContent);
          const rightValue = Number.parseFloat(right.cells[colIndex].textContent);
          const leftSortable = Number.isNaN(leftValue) ? -Infinity : leftValue;
          const rightSortable = Number.isNaN(rightValue) ? -Infinity : rightValue;
          return nextDir === "asc"
            ? leftSortable - rightSortable
            : rightSortable - leftSortable;
        }});
        rows.forEach((row) => tbody.appendChild(row));
      }});
    }});
  </script>
</body>
</html>
"""
    path.write_text(html_text, encoding="utf-8")


def load_existing_output(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    if not path.exists():
        raise FileNotFoundError(f"Existing Excel output not found: {path}")

    prices = pd.read_excel(path, sheet_name="prices_XXXUSD", index_col=0)
    prices.index = pd.to_datetime(prices.index)
    prices.index = prices.index.normalize()
    prices = prices.sort_index().groupby(level=0).last()

    metadata = pd.read_excel(path, sheet_name="metadata")
    return prices, metadata


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a static USD FX correlation dashboard from TradingView data.")
    parser.add_argument("--exchange", default=DEFAULT_EXCHANGE, help="TradingView exchange/feed. Default: FX_IDC")
    parser.add_argument("--bars", type=int, default=DEFAULT_BARS, help="Daily candles to request per symbol. Default: 250")
    parser.add_argument("--st-window", type=int, default=DEFAULT_ST_WINDOW, help="Short-term return window. Default: 30")
    parser.add_argument("--lt-window", type=int, default=DEFAULT_LT_WINDOW, help="Long-term return window. Default: 180")
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR), help="Output directory. Default: output")
    parser.add_argument(
        "--use-existing-output",
        action="store_true",
        help="Regenerate reports from the existing output Excel file instead of fetching TradingView data.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    excel_path = output_dir / "usd_fx_correlation.xlsx"
    html_path = output_dir / "usd_fx_correlation.html"

    if args.use_existing_output:
        prices, metadata = load_existing_output(excel_path)
    else:
        prices, metadata = fetch_all(USD_CURRENCIES, EXTRA_INSTRUMENTS, args.exchange, args.bars)
    log_returns = calculate_log_returns(prices)
    prices = order_frame_columns(prices)
    log_returns = order_frame_columns(log_returns)
    corr_st = order_matrix(calculate_window_correlation(log_returns, args.st_window))
    corr_lt = order_matrix(calculate_window_correlation(log_returns, args.lt_window))
    corr_diff = corr_st - corr_lt

    write_excel(
        excel_path,
        prices,
        log_returns,
        corr_st,
        corr_lt,
        corr_diff,
        metadata,
        args.st_window,
        args.lt_window,
    )
    write_html(
        html_path,
        prices,
        log_returns,
        corr_st,
        corr_lt,
        corr_diff,
        metadata,
        args.st_window,
        args.lt_window,
    )

    print(f"wrote {excel_path}")
    print(f"wrote {html_path}")


if __name__ == "__main__":
    main()
